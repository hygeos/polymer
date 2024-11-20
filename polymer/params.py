#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import print_function, division, absolute_import
import numpy as np
from os.path import join, dirname
from collections import OrderedDict
from pyhdf.SD import SD
from polymer.hico import bands_hico, K_OZ_HICO, K_NO2_HICO
from polymer import prisma
from scipy.interpolate import interp1d

# pass these parameters to polymer to obtain the quasi-same results as polymer v3.5
# polymer(<level>, <level2>, **params_v3_5)
params_v3_5 = {
    'reinit_rw_neg': True,
    'constraint_logfb': [1e-3, 0.2258, 0.9233],
    'metrics': 'polymer_3_5',
    'Rprime_consistency': False,
    'absorption': 'bricaud95_aphy',
    }


# these parameters will quasi-reproduce the VIIRS v3.5 with MM01 model
params_v3_5_VIIRS_MM01 = {
        'reinit_rw_neg': True,
        'constraint_logfb': [1e-3, 0.2258, 0.9233],
        'metrics': 'polymer_3_5',
        'Rprime_consistency': False,
        'absorption': 'bricaud95_aphy',
        'bands_corr' : [410,443,486,551,671,745,862],
        'bands_oc'   : [410,443,486,551,671,745,862],
        'water_model' : 'MM01_FOQ',
        'atm_model' : 'T0,-1,-4',
        'calib' : {  # vicarious calibration R2014.0
            410: 0.9631, 443: 1.0043,
            486: 1.0085, 551: 0.9765,
            671: 1.0204, 745: 1.0434,
            862: 1.0   , 1238:1.0   ,
            },
        'thres_Rcloud_std': 0.02,
    }

params_minabs2 = {
    # merge mineral absorption with chl
    # Note: add bands 1020 for OLCI and 1610 for MSI in bands_corr and bands_oc
    # (see params_minabs2_MSI and params_minabs2_OLCI)
    'min_abs': -2,
    'bounds': [[-2, 4], [-2, 3.5]],
    'initial_points': [
        [x, y]
        for x in [-2, -1, 0, 1]
        for y in [0, 1]
    ] + [[1, 2], [2, 2]],
}

params_minabs2_OLCI = {
    'bands_corr': [443,490,510,560,620,665,754,779,865,1020],
    'bands_oc':   [443,490,510,560,620,665,754,779,865,1020],
    **params_minabs2
}

params_minabs2_MSI = {
    'bands_corr': [443,490,560,665,705,740,783,865,1610],
    'bands_oc':   [443,490,560,665,705,740,783,865,1610],
    'weights_corr': [1,1,1,1,1,1,1,1,0.01],
    'weights_oc':   [1,1,1,1,1,1,1,1,0.01],
    **params_minabs2
}

# OLCI parameters consistent with params_v3_5_VIIRS_MM01
params_v3_5_OLCI_MM01 = {
        'reinit_rw_neg': True,
        'constraint_logfb': [1e-3, 0.2258, 0.9233],
        'metrics': 'polymer_3_5',
        'Rprime_consistency': False,
        'absorption': 'bricaud95_aphy',
        'water_model' : 'MM01_FOQ',
        'atm_model' : 'T0,-1,-4',
        }


class Params(object):
    '''
    A class to store the processing parameters
    '''
    def __init__(self, sensor, **kwargs):

        # store attributes in an OrderedDict
        self.__dict__['_odict'] = OrderedDict()

        if 'dir_base' in kwargs:
            self.dir_base = kwargs['dir_base']
        else:
            self.dir_base = dirname(dirname(__file__))
        self.sensor = sensor

        self.dir_common = join(self.dir_base, 'auxdata/common/')

        # define common parameters
        self.common(**kwargs)

        # define sensor-specific parameters
        self.sensor_specific(sensor)

        # setup custom parameters
        self.update(**kwargs)

        # finalization
        self.finalize()

    def common(self, **kwargs):
        '''
        define common parameters
        '''

        # cloud masking (negative values disactivate)
        self.thres_Rcloud = 0.2
        self.thres_Rcloud_std = 0.04  # may be overridden by sensor-specific options

        # optimization parameters
        self.force_initialization = False
        self.reinit_rw_neg = False
        self.max_iter = 100
        self.size_end_iter = 0.005
        self.metrics = 'W_dR2_norm'
        self.glint_precorrection = True
        self.external_mask = None

        # Generic look-up table
        self.lut_file = join(self.dir_base, 'auxdata/generic/LUT.hdf')

        self.thres_chi2 = 0.005

        self.partial = 0    # whether to perform partial processing
                            #   0: standard processing
                            #   1: stop before minimize
                            #   2: stop before rayleigh correction
                            #   3: stop before cloud mask
                            #   4: stop before gaseous correction
                            #   5: stop before conversion to reflectance

        # weights_corr_and weights_oc:
        # weights for atmospheric (corr) and cost function (oc) fits
        # can be:
        #  * array-like of weights, of same size as bands_corr and bands_oc
        #  * function taking bands_corr or bands_oc as input, returning array-like of weights
        #  * strings eval'd to such function
        self.weights_corr = None
        self.weights_oc = None

        self.atm_model = 'T0,-1,Rmol'

        # water reflectance normalization
        #   * no geometry nor wavelength normalization (0)
        #   * apply normalization of the water reflectance for
        #     bidirectional effects (BRDF) to the nadir-nadir geometry (1)
        #   * apply wavelength normalization for MERIS and OLCI (2)
        #   * apply both (3)
        self.normalize = 3

        self.Rprime_consistency = True

        # water model
        # PR05: based on Park and Ruddick, 2005
        # MM01: based on Morel and Maritorena, 2001
        # MM01_FOQ: MM01 including directional reflectances using f/Q tables
        if 'water_model' in kwargs:
            self.water_model = kwargs['water_model']
        else:
            self.water_model = 'PR05'

        if self.water_model == 'PR05':
            self.initial_point_1 = [-1, 0]
            self.initial_point_2 = [1, 1]
            self.initial_points = [[]]
            self.initial_step = [0.2, 0.2]
            self.firstguess_method = 1
            self.bounds = [[-2, 2], [-3, 3]]

            # Constraint on bbs: amplitude, sigma(chl=0.01), sigma(chl=0.1)
            # (disactivate with amplitude == 0)
            # NOTE: amplitude changed from 1e-3 to 1e-4 since sumsq is now
            # divided by the sum of weights (inter-sensor consistency)
            self.constraint_logfb = [1e-4, 0.2258, 0.9233]

            # PR05 model only
            self.bbopt = 0      # particle backscattering model
            self.min_abs = 0           # 0: don't include particle absorption
                                       # 1: include mineral absorption (data from HZG)
                                       # 2: include NAP absorption (Babin2003)
            self.absorption = 'bricaud98_aphy'

        elif self.water_model.startswith('MM01'):
            self.initial_point_1 = [-1, 0]
            self.initial_point_2 = [-1, 0]
            self.initial_step = [0.05, 0.0005]
            self.bounds = [[-2, 2], [-0.005, 0.1]]

            # Constraint on bbs: amplitude, sigma(chl=0.01), sigma(chl=0.1)
            # (disactivate with amplitude == 0)
            self.constraint_logfb = [1e-3, 0.0001, 0.005]
            raise NotImplementedError


        # no2 absorption data
        self.no2_climatology = join(self.dir_base, 'auxdata/common/no2_climatology.hdf')
        self.no2_frac200m  = join(self.dir_base, 'auxdata/common/trop_f_no2_200m.hdf')

        self.multiprocessing = 0 # 0: single thread
                                 # N != 0: multiprocessing, with:
                                 # N < 0: use as many threads as there are CPUs

        # Digital Elevation Model (DEM)
        # can be:
        #     * a constant (use this constant altitude for the whole scene)
        #     * a DEM_SRTM object
        self.altitude = 0

        # BITMASK
        # (see common.py for bitmask definition)
        # no product (NaN) in case of...
        self.BITMASK_INVALID = 1+2+4+32+512

        # recommended pixek rejection test: (bitmask & BITMASK_REJECT) != 0
        self.BITMASK_REJECT = 1023

        self.verbose = True

        self.dbg_pt = [-1, -1]

        # uncertainties
        self.uncertainties = 0

    def sensor_specific(self, sensor):
        '''
        define sensor-specific default parameters
        '''
        if sensor in ['MERIS', 'MERIS_FR', 'MERIS_RR']:
            self.defaults_meris()
        elif sensor == 'MSI':
            self.defaults_msi()
        elif sensor == 'OLCI':
            self.defaults_olci()
        elif sensor in ['VIIRS', 'VIIRSN']:
            self.defaults_viirsn()
        elif sensor == 'VIIRSJ1':
            self.defaults_viirsj1()
        elif sensor == 'MODIS':
            self.defaults_modis()
        elif sensor == 'SeaWiFS':
            self.defaults_seawifs()
        elif sensor == 'HICO':
            self.defaults_hico()
        elif sensor == 'OLI':
            self.defaults_oli()
        elif sensor == 'PRISMA':
            self.defaults_prisma()
        elif sensor == 'GENERIC':
            self.defaults_generic()
        else:
            raise Exception('Params.sensor_specific: invalid sensor "{}"'.format(sensor))

    def defaults_generic(self):
        '''
        sensor_specific parameters should be provided by the user
        '''

        self.bands_corr = []
        self.bands_oc =   []
        self.bands_rw =   []
        self.band_cloudmask = -999
        self.calib = {}
        self.K_OZ = {}
        self.K_NO2 = {}

    def defaults_meris(self):
        '''
        define default parameters for ENVISAT/MERIS
        '''

        self.bands_corr = [412,443,490,510,560,620,665,        754,    779,865]
        self.bands_oc =   [412,443,490,510,560,620,665,        754,    779,865]
        self.bands_rw =   [412,443,490,510,560,620,665,681,709,754,    779,865]

        self.band_cloudmask = 865

        # self.calib = {
                    # 412: 1.0, 443: 1.0, 490: 1.0, 510: 1.0,
                    # 560: 1.0, 620: 1.0, 665: 1.0, 681: 1.0,
                    # 709: 1.0, 754: 1.0, 760: 1.0, 779: 1.0,
                    # 865: 1.0, 885: 1.0, 900: 1.0,
                # }

        self.calib = {  # OC-CCI VC 20161215 (NCEP)
                412: 0.997727, 443: 0.999919,
                490: 0.996547, 510: 0.999385,
                560: 0.996184, 620: 1.002931,
                665: 0.999352, 754: 1.000000,
                681: 1.000000, 709: 1.000000,
                760: 1.000000, 779: 1.000000,
                865: 1.000000, 885: 1.000000,
                900: 1.000000,
                }
        # self.calib = {  # OC-CCI VC 20161017 (ERA)
                # 412: 0.996985, 443: 0.998956,
                # 490: 0.995681, 510: 0.999163,
                # 560: 0.997005, 620: 1.004477,
                # 665: 0.999794, 754: 1.000000,
                # 779: 1.000000, 865: 1.000000,
                # }

        self.K_OZ = {
                    412: 0.000301800 , 443: 0.00327200 ,
                    490: 0.0211900   , 510: 0.0419600  ,
                    560: 0.104100    , 620: 0.109100   ,
                    665: 0.0511500   , 681: 0.0359600  ,
                    709: 0.0196800   , 754: 0.00955800 ,
                    760: 0.00730400  , 779: 0.00769300 ,
                    865: 0.00219300  , 885: 0.00121100 ,
                    900: 0.00151600  ,
                }

        self.K_NO2 = {
                412: 6.074E-19 , 443: 4.907E-19,
                490: 2.916E-19 , 510: 2.218E-19,
                560: 7.338E-20 , 620: 2.823E-20,
                665: 6.626E-21 , 681: 6.285E-21,
                709: 4.950E-21 , 754: 1.384E-21,
                760: 4.717E-22 , 779: 3.692E-22,
                865: 2.885E-23 , 885: 4.551E-23,
                900: 5.522E-23 ,
                }

    def defaults_olci(self):
        '''
        define default parameters for Sentinel-3/OLCI
        '''

        self.bands_corr = [        443,490,510,560,620,665,            754,779,865]
        self.bands_oc   = [        443,490,510,560,620,665,            754,779,865]
        self.bands_rw   = [400,412,443,490,510,560,620,665,674,681,709,754,779,865,885,1020]

        self.band_cloudmask = 865

        # self.calib = {   # SVC Constant Aug18 (not valid anymore with band configuration without 412nm)
                # 400 : 1.0  , 412 : 0.997,
                # 443 : 0.997, 490 : 0.989,
                # 510 : 0.993, 560 : 0.998,
                # 620 : 1.0  , 665 : 1.0,
                # 674 : 1.0  , 681 : 1.0,
                # 709 : 1.0  , 754 : 1.0,
                # 760 : 1.0  , 764 : 1.0,
                # 767 : 1.0  , 779 : 1.0,
                # 865 : 1.0  , 885 : 1.0,
                # 900 : 1.0  , 940 : 1.0,
                # 1020: 1.0  , 1375: 1.0,
                # 1610: 1.0  , 2250: 1.0,
                # }

        self.calib = {  # SVC 01/08/19
            400 : 1.000000, 412 : 1.000000,
            443 : 0.999210, 490 : 0.984674,
            510 : 0.986406, 560 : 0.988970,
            620 : 0.991810, 665 : 0.987590,
            674 : 1.000000, 681 : 1.000000,
            709 : 1.000000, 754 : 1.000000,
            760 : 1.000000, 764 : 1.000000,
            767 : 1.000000, 779 : 1.000000,
            865 : 1.000000, 885 : 1.000000,
            900 : 1.000000, 940 : 1.000000,
            1020: 1.000000,
        }

        # from SeaDAS v7.3.2
        self.K_OZ = {
                400 : 2.985E-06, 412 : 2.341E-04,
                443 : 2.897E-03, 490 : 2.066E-02,
                510 : 4.129E-02, 560 : 1.058E-01,
                620 : 1.085E-01, 665 : 5.005E-02,
                674 : 4.095E-02, 681 : 3.507E-02,
                709 : 1.887E-02, 754 : 8.743E-03,
                760 : 6.713E-03, 764 : 6.916E-03,
                767 : 6.754E-03, 779 : 7.700E-03,
                865 : 2.156E-03, 885 : 1.226E-03,
                900 : 1.513E-03, 940 : 7.120E-04,
                1020: 8.448E-05,
                }

        # from SeaDAS v7.3.2
        self.K_NO2 = {
                400 : 6.175E-19, 412 : 6.083E-19,
                443 : 4.907E-19, 490 : 2.933E-19,
                510 : 2.187E-19, 560 : 7.363E-20,
                620 : 2.818E-20, 665 : 6.645E-21,
                674 : 1.014E-20, 681 : 6.313E-21,
                709 : 4.938E-21, 754 : 1.379E-21,
                760 : 4.472E-22, 764 : 6.270E-22,
                767 : 5.325E-22, 779 : 3.691E-22,
                865 : 2.868E-23, 885 : 4.617E-23,
                900 : 5.512E-23, 940 : 3.167E-24,
                1020: 0.000E+00,
                }

    def defaults_msi(self):

        self.bands_corr = [443,490,560,665,705,740,783,    865,                  ]
        self.bands_oc   = [443,490,560,665,705,740,783,    865,                  ]
        self.bands_rw   = [443,490,560,665,705,740,783,842,865,         1610,    ]

        self.band_cloudmask = 865

        self.calib = {
                443 : 1.0,
                490 : 1.0,
                560 : 1.0,
                665 : 1.0,
                705 : 1.0,
                740 : 1.0,
                783 : 1.0,
                842 : 1.0,
                865 : 1.0,
                945 : 1.0,
                1375: 1.0,
                1610: 1.0,
                2190: 1.0,
                }

        self.K_OZ = {   # FIXME: rough values from OLCI
                443 : 2.897E-03,
                490 : 2.066E-02,
                560 : 1.058E-01,
                665 : 5.005E-02,
                705 : 1.887E-02,
                740 : 8.743E-03,
                783 : 7.700E-03,
                842 : 0.,
                865 : 2.156E-03,
                945 : 0.,
                1375: 0.,
                1610: 0.,
                2190: 0.,
                }

        self.K_NO2 = {   # FIXME
                443 : 0.,
                490 : 0.,
                560 : 0.,
                665 : 0.,
                705 : 0.,
                740 : 0.,
                783 : 0.,
                842 : 0.,
                865 : 0.,
                945 : 0.,
                1375: 0.,
                1610: 0.,
                2190: 0.,
                }

    def defaults_viirsn(self):

        self.bands_corr = [    443,486,551,671,745,862               ]
        self.bands_oc   = [    443,486,551,671,745,862               ]
        self.bands_rw   = [410,443,486,551,671,745,862               ]

        self.band_cloudmask = 862

        self.thres_Rcloud_std = 0.005

        # self.calib = {  # vicarious calibration R2014.0
                # 410: 0.9631, 443: 1.0043,
                # 486: 1.0085, 551: 0.9765,
                # 671: 1.0204, 745: 1.0434,
                # 862: 1.0   , 1238:1.0   ,
                # }

        # self.calib = {  # OC-CCI VC 20161017
        #         410:0.963100, 443:1.002699,
        #         486:1.002965, 551:0.966786,
        #         671:1.012163, 745:1.043400,
        #         862:1.000000,
        #         }

        # self.calib = {  # standard vicarious valibration R2018
        #         410 : 0.9752, 443 : 0.9579,
        #         486 : 0.9874, 551 : 0.9824,
        #         671 : 0.9918, 745 : 0.9922,
        #         862 : 1.0, 1238: 1.0,
        #         1601: 1.0, 2257: 1.0,
        #     }

        self.calib = {  # OC-CCI SVC 20190430 (ERA) for NASA R2018 (L1C generated with g=1)
                410: 0.975200, 443: 0.955771,
                486: 0.983928, 551: 0.978784,
                671: 0.991800, 745: 0.992200,
                862: 1.000000, }

        self.K_OZ = {  # from SeaDAS
                410: 5.827E-04,
                443: 3.079E-03,
                486: 1.967E-02,
                551: 8.934E-02,
                671: 4.427E-02,
                745: 1.122E-02,
                862: 2.274E-03,
                1238:0.
                }
        self.K_NO2 = {  # from SeaDAS
                410: 5.914E-19,
                443: 5.013E-19,
                486: 3.004E-19,
                551: 1.050E-19,
                671: 1.080E-20,
                745: 2.795E-21,
                862: 3.109E-22,
                1238:0.000E+00,
                }

    def defaults_viirsj1(self):

        self.bands_corr = [    445,489,556,667,746,868               ]
        self.bands_oc   = [    445,489,556,667,746,868               ]
        self.bands_rw   = [411,445,489,556,667,746,868               ]

        self.band_cloudmask = 868

        self.thres_Rcloud_std = 0.005

        # SeaDAS gains 7.5.3
        # self.calib = {
        #     411 : 1.0017, 445 : 1.0160, 489 : 1.0222,
        #     556 : 1.0115, 667 : 1.0088, 746 : 1.0099,
        #     868 : 1.0, 1238: 1.0, 1604: 1.0, 2258: 1.0,
        #     }

        self.calib = None

        self.K_OZ = {  # from SeaDAS
                411 : 2.618E-04,
                445 : 3.220E-03,
                489 : 2.129E-02,
                556 : 9.965E-02,
                667 : 4.771E-02,
                746 : 1.083E-02,
                868 : 2.062E-03,
                1238: 3.897E-14,
                1604: 2.357E-14,
                2258: 1.388E-13,
                }
        self.K_NO2 = {  # from SeaDAS
                411 : 6.006E-19,
                445 : 4.932E-19,
                489 : 2.886E-19,
                556 : 8.767E-20,
                667 : 8.008E-21,
                746 : 1.820E-21,
                868 : 6.223E-23,
                1238: 0.000E+00,
                1604: 0.000E+00,
                2258: 0.000E+00,
                }

    def defaults_seawifs(self):
        self.bands_corr = [412,443,490,510,555,670,    865]
        self.bands_oc   = [412,443,490,510,555,670,    865]
        self.bands_rw   = [412,443,490,510,555,670,    865]

        self.band_cloudmask = 865

        # self.calib = {  # OC-CCI VC 20161017
                # 412: 0.993524, 443: 0.989207,
                # 490: 0.988410, 510: 0.986153,
                # 555: 1.000000, 670: 1.000000,
                # 865: 1.000000,
                # }

        self.calib = {  # OC-CCI SVC 20190417 (ERA) for NASA R2018 (L1C generated with g=1)
                412: 0.999438, 443: 0.995541,
                490: 0.993768, 510: 0.992603,
                555: 1.000000, 670: 1.000000,
                865: 1.000000,
                }

        self.K_OZ = {  # from SeaDAS
                412: 4.114E-04,
                443: 3.162E-03,
                490: 2.346E-02,
                510: 4.094E-02,
                555: 9.568E-02,
                670: 4.649E-02,
                765: 8.141E-03,
                865: 3.331E-03,
                }
        self.K_NO2 = {  # from SeaDAS
                412: 6.004E-19,
                443: 4.963E-19,
                490: 2.746E-19,
                510: 2.081E-19,
                555: 9.411E-20,
                670: 9.234E-21,
                765: 1.078E-21,
                865: 1.942E-21,
                }

    def defaults_modis(self):
        self.bands_corr = [412,443,    488,531,547,        667,678,748,    869,    ]
        self.bands_oc   = [412,443,    488,531,547,        667,678,748,    869,    ]
        self.bands_rw   = [412,443,    488,531,547,        667,678,748,    869,    ]

        self.band_cloudmask = 869

        # self.calib = {  # OC_CCI VICARIOUS CALIBRATION 2015 (VIC2, ERA-INTERIM)
                # 412 : 0.995, 443 : 1.001,
                # 469 : 1.0  , 488 : 1.000,
                # 531 : 1.001, 547 : 1.000,
                # 555 : 1.0  , 645 : 1.0  ,
                # 667 : 0.992, 678 : 1.0  ,
                # 748 : 1.0  , 859 : 1.0  ,
                # 869 : 1.0  , 1240: 1.0  ,
                # }

        # self.calib = {  # OC-CCI VC 20161215 (NCEP)
            # 412:1.019888, 443:1.025089,
            # 488:1.020164, 531:1.018262,
            # 547:1.017205, 667:1.007744,
            # 678:1.000000, 748:1.000000,
            # 869:1.000000,
            # }
        # self.calib = {  # OC-CCI VC 20161017 (ERA)
            # 412:1.014446, 443:1.019815,
            # 488:1.015669, 531:1.015656,
            # 547:1.015389, 667:1.007577,
            # 678:1.000000, 748:1.000000,
            # 869:1.000000,
            # }

        self.calib = { # OC-CCI SVC 20190425 (ERA) for NASA R2018 (L1C generated with g=1)
            412: 0.998926, 443: 1.002921,
            488: 0.998408, 531: 0.998077,
            547: 0.997570, 667: 1.000000,
            678: 1.000000, 748: 1.000000,
            869: 1.000000,
            469: 1., 555: 1., 645: 1.,
            859: 1., 1240:1.,
            }

        self.K_OZ = {  # from SeaDAS
                412 :1.987E-03,
                443 :3.189E-03,
                469 :8.745E-03,
                488 :2.032E-02,
                531 :6.838E-02,
                547 :8.622E-02,
                555 :9.553E-02,
                645 :7.382E-02,
                667 :4.890E-02,
                678 :3.787E-02,
                748 :1.235E-02,
                859 :2.347E-03,
                869 :1.936E-03,
                1240:0.000E+00,
                }
        self.K_NO2 = {  # from SeaDAS
                412 :5.814E-19,
                443 :4.985E-19,
                469 :3.938E-19,
                488 :2.878E-19,
                531 :1.525E-19,
                547 :1.194E-19,
                555 :9.445E-20,
                645 :1.382E-20,
                667 :7.065E-21,
                678 :8.304E-21,
                748 :2.157E-21,
                859 :6.212E-23,
                869 :7.872E-23,
                1240:0.000E+00,
                }


    def defaults_hico(self):
        bands_to_use = [
                                                                        410,
            416,  421,  427,  433,  438,  444,  450,  456,  461,  467,  473,
            479,  484,  490,  496,  501,  507,  513,  519,  524,  530,  536,
            542,  547,  553,  559,  564,  570,  576,  582,  587,  593,  599,
            605,  610,  616,  622,  627,  633,  639,  645,  650,  656,  662,
                        742,  748,  753,              771,  776,
            ]
        self.bands_corr = bands_to_use
        self.bands_oc = bands_to_use
        self.bands_rw = [                                               410,
            416,  421,  427,  433,  438,  444,  450,  456,  461,  467,  473,
            479,  484,  490,  496,  501,  507,  513,  519,  524,  530,  536,
            542,  547,  553,  559,  564,  570,  576,  582,  587,  593,  599,
            605,  610,  616,  622,  627,  633,  639,  645,  650,  656,  662,
            668,  673,  679,  685,  690,  696,  702,  708,  713,  719,  725,
            731,  736,  742,  748,  753,  759,  765,  771,  776,  782,  788,
            794,  799,  805,  811,  816,  822,  828,  834,  839,  845,  851,
            857,  862,  868,
            ]

        self.calib = {   # from Ibrahim et al, 2017,
                         # Atmospheric correction for hyperspectral ocean color retrieval with
                         # application to the Hyperspectral Imager for the Coastal Ocean (HICO)
            404: 1.042, 410: 1.012, 416: 1.009, 421: 1.025, 427: 1.034, 433: 1.017, 438: 1.019, 444: 1.028,
            450: 1.035, 456: 1.046, 461: 1.051, 467: 1.044, 473: 1.041, 479: 1.041, 484: 1.023, 490: 1.013,
            496: 1.020, 501: 1.013, 507: 1.013, 513: 1.000, 519: 0.991, 524: 0.993, 530: 0.998, 536: 1.006,
            542: 1.014, 547: 1.020, 553: 1.021, 559: 1.015, 564: 1.016, 570: 1.028, 576: 1.038, 582: 1.039,
            587: 1.036, 593: 1.033, 599: 1.038, 605: 1.038, 610: 1.038, 616: 1.038, 622: 1.050, 627: 1.060,
            633: 1.063, 639: 1.063, 645: 1.058, 650: 1.057, 656: 1.046, 662: 1.044, 668: 1.043, 673: 1.041,
            679: 1.041, 685: 1.023, 690: 1.026, 696: 1.063, 702: 1.076, 708: 1.072, 713: 1.069, 719: 1.058,
            725: 1.059, 731: 1.066, 736: 1.059, 742: 1.052, 748: 1.038, 753: 1.022, 759: 0.986, 765: 0.973,
            771: 0.995, 776: 1.000, 782: 1.005, 788: 1.000, 794: 0.992, 799: 0.981,
            # complete with ones
            805: 1.,811: 1.,816: 1.,822: 1.,828: 1.,834: 1.,839: 1.,845: 1.,851: 1., 857: 1.,862: 1.,868: 1.,
            }

        self.K_OZ = K_OZ_HICO
        self.K_NO2 = K_NO2_HICO

        self.band_cloudmask = 862


    def defaults_prisma(self):
        self.bands_corr = [
                                          453, 460, 468, 475,
            482, 489, 497, 504, 512, 519, 527, 535, 542, 550,
            559, 567, 575, 583, 592, 601, 609, 618, 627, 636,
            645, 655, 664,
                                785, 796, 806,
            849, 859, 870, 881,
        ]
        self.bands_oc = self.bands_corr
        self.bands_rw = [
            406, 415, 423, 431, 438, 446, 453, 460, 468, 475,
            482, 489, 497, 504, 512, 519, 527, 535, 542, 550,
            559, 567, 575, 583, 592, 601, 609, 618, 627, 636,
            645, 655, 664, 674, 684, 694, 703, 713, 723, 733,
            744, 754, 764, 775, 785, 796, 806, 817, 827, 838,
            849, 859, 870, 881, 892, 902, 913, 923, 934, 944,
            956, 967, 977
        ]

        self.K_OZ = prisma.K_OZ
        self.K_NO2 = prisma.K_NO2
        self.calib = {k: 1. for k in prisma.bands}
        self.band_cloudmask = 859
        self.thres_Rcloud = 0.1

    def defaults_oli(self):
        '''
        Landsat8/OLI defaults
        '''
        # QV 2024-07-25 use openpyxl instead of xlrd
        from openpyxl import load_workbook
        import pandas as pd

        self.bands_corr = [440, 480, 560, 655, 865      ]
        self.bands_oc   = [440, 480, 560, 655, 865      ]
        self.bands_rw   = [440, 480, 560, 655, 865      ]

        self.calib = {
                440: 1.,
                480: 1.,
                560: 1.,
                655: 1.,
                865: 1.,
                1610: 1.,
                2200: 1.,
                }


        # Ozone optical depth for 1000 DU
        self.K_OZ = OrderedDict()
        srf_file = join(self.dir_base, 'auxdata', 'oli',
                        'Ball_BA_RSR.v1.2.xlsx')
        wb = load_workbook(filename = srf_file)

        solar_spectrum_file = join(self.dir_common, 'SOLAR_SPECTRUM_WMO_86')
        solar_data = pd.read_csv(solar_spectrum_file, sep=' ')

        k_oz_data = pd.read_csv(join(self.dir_common, 'k_oz.csv'),
                                skiprows=1)
        k_oz = interp1d(k_oz_data.wavelength, k_oz_data.K_OZ,
                        bounds_error=False, fill_value=0.)
        F0 = interp1d(solar_data['lambda(nm)'], solar_data['Sl(W.m-2.nm-1)'])

        for b, bname in [(440, 'CoastalAerosol'),
                         (480, 'Blue'),
                         (560, 'Green'),
                         (655, 'Red'),
                         (865, 'NIR'),
                         (1610, 'SWIR1'),
                         (2200, 'SWIR2')]:
            sh = wb[bname]

            wav, srf = [], []

            i=1
            while i<wb[bname].max_row:
                i += 1
                if sh.cell(i, 1).value is None: break
                wav.append(sh.cell(i, 1).value)
                srf.append(sh.cell(i, 2).value)

            wav, srf = np.array(wav), np.array(srf)

            self.K_OZ[b] = np.trapz(k_oz(wav)*srf*F0(wav))/np.trapz(srf*F0(wav))

        self.K_NO2 = {
                440: 0.,
                480: 0.,
                560: 0.,
                655: 0.,
                865: 0.,
                1610: 0.,
                2200: 0.,
                }

        self.band_cloudmask = 865

    def bands_read(self):
        assert (np.diff(self.bands_corr) > 0).all()
        assert (np.diff(self.bands_oc) > 0).all()
        assert (np.diff(self.bands_rw) > 0).all()
        bands_read = set(self.bands_corr)
        bands_read = bands_read.union(self.bands_oc)
        bands_read = bands_read.union(self.bands_rw)
        return sorted(bands_read)

    def preprocess(self, l1):
        '''
        This method is executed after params initialization
        '''
        #
        # initialize external mask
        #
        if self.external_mask is not None:
            if isinstance(self.external_mask, str) and self.external_mask.endswith(".nc"):
                import xarray as xr
                with xr.open_dataset(self.external_mask) as ds:
                    self.external_mask = ds["mask"].values
            elif isinstance(self.external_mask, str):
                # read external mask
                hdf = SD(self.external_mask)
                self.external_mask = hdf.select('mask').get()
            elif isinstance(self.external_mask, np.ndarray):
                pass
            else:
                raise Exception('external_mask should be a hdf filename containing a mask dataset, or a numpy array, or None')

            # check external mask size
            assert (l1.height, l1.width) == self.external_mask.shape, \
                    'Error, product shape is {} but external mask shape is {}'.format( (l1.width, l1.height), self.external_mask.shape)


    def print_info(self):
        print(self.__class__)
        for k, v in self.items():
            print('*', k,':', v)

    def update(self, **kwargs):
        self.__dict__['_odict'].update(kwargs)

    def __getattr__(self, key):
        return self.__dict__['_odict'][key]

    def __setattr__(self, key, value):
        self.__dict__['_odict'][key] = value

    def items(self):
        return self.__dict__['_odict'].items()

    def __getstate__(self):
        # make this class picklable
        return dict(self.__dict__['_odict'])

    def __setstate__(self, state):
        # make this class picklable
        self.__dict__['_odict'] = state

    def finalize(self):

        if isinstance(self.weights_corr, str):
            self.weights_corr = eval(self.weights_corr)
        if hasattr(self.weights_corr, '__call__'):
            self.weights_corr = self.weights_corr(self.bands_corr)

        if isinstance(self.weights_oc, str):
            self.weights_oc = eval(self.weights_oc)
        if hasattr(self.weights_oc, '__call__'):
            self.weights_oc = self.weights_oc(self.bands_oc)

        # number of terms in the model
        self.Ncoef = self.atm_model.count(',')+1
